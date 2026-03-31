import csv
import io
import os
import datetime
from flask import Blueprint, request, jsonify, send_file, has_request_context
from bson import ObjectId
from urllib.parse import unquote
from ..db import get_db, get_nursing_employees_db, ensure_nursing_dept_db

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".ods"}


def _file_extension(filename: str) -> str:
    return os.path.splitext(filename.lower())[1]


def _rows_from_xlsx(file_bytes: bytes) -> list[list[str]]:
    """Parse .xlsx using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(c for c in cells):
            rows.append(cells)
    return rows


def _rows_from_xls(file_bytes: bytes) -> list[list[str]]:
    """Parse legacy .xls using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)
    rows = []
    for i in range(ws.nrows):
        cells = [str(ws.cell_value(i, j)).strip() for j in range(ws.ncols)]
        # xlrd returns floats for numeric cells; strip trailing .0
        cells = [c[:-2] if c.endswith(".0") else c for c in cells]
        if any(c for c in cells):
            rows.append(cells)
    return rows


def _rows_from_ods(file_bytes: bytes) -> list[list[str]]:
    """Parse .ods (LibreOffice Calc) using odfpy."""
    from odf.opendocument import load as ods_load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = ods_load(io.BytesIO(file_bytes))
    spreadsheet = doc.spreadsheet
    rows = []
    for sheet in spreadsheet.getElementsByType(Table):
        for row_el in sheet.getElementsByType(TableRow):
            cells = []
            for cell in row_el.getElementsByType(TableCell):
                paragraphs = cell.getElementsByType(P)
                text = (
                    " ".join(str(p) for p in paragraphs).strip() if paragraphs else ""
                )
                repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                cells.extend([text] * repeat)
            # Trim trailing empty cells
            while cells and not cells[-1]:
                cells.pop()
            if any(c for c in cells):
                rows.append(cells)
        break  # first sheet only
    return rows


def _rows_from_csv(file_bytes: bytes) -> list[list[str]]:
    """Parse .csv (handles BOM from Excel)."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return [r for r in reader if any(cell.strip() for cell in r)]


def _get_rows(file_bytes: bytes, ext: str) -> list[list[str]]:
    """Dispatch to the correct parser based on file extension."""
    if ext == ".xlsx":
        return _rows_from_xlsx(file_bytes)
    elif ext == ".xls":
        return _rows_from_xls(file_bytes)
    elif ext == ".ods":
        return _rows_from_ods(file_bytes)
    else:
        return _rows_from_csv(file_bytes)


_DEFAULTS_CSV = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "default_shift_types.csv"
)

employees_bp = Blueprint("employees", __name__)


def _serialize(doc):
    doc["id"] = str(doc.pop("_id"))
    return doc


TICKED_VALUES = {"v", "✓", "✔", "x", "yes", "כן", "1", "true"}

def col_attr(i):
    return f"col_{i}"


HOME_DEPT_HEADER = "מחלקת אם"
HOME_DEPT_HEADER_ALT = "מחלקה"


def _parse_rows(rows: list[list[str]]) -> tuple[list[str], list[dict]]:
    """
    Parse pre-extracted rows into employees and column headers.

    Row 0  = headers (column names).
    Rows 1+ = employees: col 0 = name, cols 1–N = attribute columns.
    A cell is considered "ticked" when its value (case-insensitive) is one of:
      V, v, ✓, ✔, X, x, yes, כן, 1, true
    A column named "מחלקת אם" is treated as a free-text home_department field
    (not a boolean attribute).
    Returns (column_headers, employees_list).
    """
    if not rows:
        return [], []

    headers = [h.strip() for h in rows[0]]

    # Find the home_department column index (1-based), if present
    home_dept_col = next(
        (i for i, h in enumerate(headers) if h in (HOME_DEPT_HEADER, HOME_DEPT_HEADER_ALT)), None
    )

    # Build attribute column headers, skipping the home_department column
    col_headers = {}
    attr_col_num = 1
    for i in range(1, len(headers)):
        if i == home_dept_col:
            continue
        col_headers[attr_col_num] = headers[i]
        attr_col_num += 1

    employees = []
    for row in rows[1:]:
        name = row[0].strip() if row else ""
        if not name:
            continue

        # Extract home_department if column exists
        home_department = None
        if home_dept_col is not None and home_dept_col < len(row):
            val = row[home_dept_col].strip()
            if val:
                home_department = val

        # Parse boolean attribute columns
        ticked_attributes = []
        attr_col_num = 1
        for raw_col in range(1, len(headers)):
            if raw_col == home_dept_col:
                continue
            cell = row[raw_col].strip() if raw_col < len(row) else ""
            if cell.lower() in TICKED_VALUES:
                ticked_attributes.append(col_attr(attr_col_num))
            attr_col_num += 1

        emp = {
            "name": name,
            "attributes": ticked_attributes,
            "col_headers": col_headers,
        }
        if home_department:
            emp["home_department"] = home_department
        employees.append(emp)

    return list(col_headers.values()), employees


# ---------------------------------------------------------------------------
# Employees — list / delete
# ---------------------------------------------------------------------------


@employees_bp.get("/employees")
def list_employees():
    db = get_nursing_employees_db()
    employees = []
    for emp in db.employees.find():
        emp["id"] = str(emp.pop("_id"))
        employees.append(emp)
    return jsonify(employees)


@employees_bp.get("/employees/export")
def export_employees():
    """Export all employees to an XLSX file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_nursing_employees_db()
    config = db.config.find_one({"key": "csv_column_headers"})
    col_headers = config["headers"] if config else []

    ids_param = request.args.get("ids", "").strip()
    if ids_param:
        id_list = [oid for s in ids_param.split(",") if (oid := s.strip())]
        object_ids = []
        for oid in id_list:
            try:
                object_ids.append(ObjectId(oid))
            except Exception:
                pass
        employees = list(db.employees.find({"_id": {"$in": object_ids}}))
        # Preserve filter order
        order = {str(e["_id"]): i for i, e in enumerate(employees)}
        employees.sort(key=lambda e: order.get(str(e["_id"]), 0))
    else:
        employees = list(db.employees.find())

    # Detect if any employee has home_department
    has_dept = any(e.get("home_department") for e in employees)

    # Build header row
    headers = ["שם"]
    if has_dept:
        headers.append("מחלקת אם")
    headers += col_headers

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "עובדים"
    ws.sheet_view.rightToLeft = True

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    dept_fill_h = PatternFill("solid", fgColor="065F46")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    even_fill   = PatternFill("solid", fgColor="F8FAFC")
    odd_fill    = PatternFill("solid", fgColor="FFFFFF")
    dept_fill   = PatternFill("solid", fgColor="ECFDF5")
    v_fill      = PatternFill("solid", fgColor="DCFCE7")
    v_font      = Font(bold=True, color="166534", name="Arial", size=10)
    name_font   = Font(bold=True, name="Arial", size=10)
    dept_font   = Font(italic=True, color="065F46", name="Arial", size=10)
    plain_font  = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = dept_fill_h if h == "מחלקת אם" else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Write employees
    for row_idx, emp in enumerate(employees, 2):
        row_fill = even_fill if row_idx % 2 == 0 else odd_fill
        attrs = emp.get("attributes", [])
        col = 1

        # Name
        cell = ws.cell(row=row_idx, column=col, value=emp.get("name", ""))
        cell.font = name_font; cell.fill = row_fill; cell.alignment = right; cell.border = border
        col += 1

        # Home department
        if has_dept:
            dept_val = emp.get("home_department") or ""
            cell = ws.cell(row=row_idx, column=col, value=dept_val)
            cell.font = dept_font; cell.fill = dept_fill if dept_val else row_fill
            cell.alignment = right; cell.border = border
            col += 1

        # Attribute columns
        for attr_idx, _ in enumerate(col_headers, 1):
            val = "V" if f"col_{attr_idx}" in attrs else ""
            cell = ws.cell(row=row_idx, column=col, value=val)
            if val == "V":
                cell.font = v_font; cell.fill = v_fill
            else:
                cell.font = plain_font; cell.fill = row_fill
            cell.alignment = center; cell.border = border
            col += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    if has_dept:
        ws.column_dimensions["B"].width = 14
    for i in range(2 + (1 if has_dept else 0), len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.row_dimensions[1].height = 36
    for r in range(2, len(employees) + 2):
        ws.row_dimensions[r].height = 22
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="employees.xlsx",
    )


@employees_bp.put("/employees/<emp_id>")
def update_employee(emp_id):
    """Update an employee's name, attributes, and/or active status."""
    db = get_nursing_employees_db()
    data = request.get_json()
    patch = {}
    if "name" in data:
        name = data["name"].strip()
        if name:
            patch["name"] = name
    if "attributes" in data:
        patch["attributes"] = [a.strip() for a in data["attributes"] if a.strip()]
    if "active" in data:
        patch["active"] = bool(data["active"])
        if not patch["active"]:
            # Record when the employee was deactivated
            patch.setdefault("inactive_since", datetime.date.today().isoformat())
        else:
            # Clear deactivation metadata when reactivating
            patch["inactive_reason"] = None
            patch["inactive_since"] = None
    if "inactive_reason" in data:
        patch["inactive_reason"] = (data["inactive_reason"] or "").strip() or None
    if "max_shifts_per_week" in data:
        val = data["max_shifts_per_week"]
        if val is not None:
            try:
                patch["max_shifts_per_week"] = max(1, int(val))
            except (TypeError, ValueError):
                pass
        else:
            patch["max_shifts_per_week"] = None
    if "home_department" in data:
        patch["home_department"] = (data["home_department"] or "").strip() or None
    if not patch:
        emp = db.employees.find_one({"_id": ObjectId(emp_id)})
        return jsonify(_serialize(emp))
    db.employees.update_one({"_id": ObjectId(emp_id)}, {"$set": patch})

    # Auto-provision department DB when home_department changes (nursing mode)
    new_dept = patch.get("home_department")
    if new_dept and has_request_context():
        raw = request.headers.get("X-App-Mode", "").strip()
        mode = unquote(raw).replace(" ", "_").replace("-", "_") if raw else ""
        if mode.startswith("nursing"):
            from .departments import auto_add_department
            auto_add_department(new_dept)
            ensure_nursing_dept_db(new_dept)

    emp = db.employees.find_one({"_id": ObjectId(emp_id)})
    return jsonify(_serialize(emp))


@employees_bp.delete("/employees/<emp_id>")
def delete_employee(emp_id):
    db = get_nursing_employees_db()
    db.employees.delete_one({"_id": ObjectId(emp_id)})
    return jsonify({"ok": True})


@employees_bp.delete("/employees")
def clear_employees():
    """Delete all employees (used before re-importing)."""
    db = get_nursing_employees_db()
    db.employees.delete_many({})
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# CSV import
# ---------------------------------------------------------------------------


@employees_bp.post("/employees/import")
def import_employees():
    """
    Accepts a multipart/form-data upload with field name 'file'.
    Replaces all existing employees with the parsed CSV data.
    Also stores column headers in the db.config collection for display.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    ext = _file_extension(file.filename or "")
    if not file.filename or ext not in ALLOWED_EXTENSIONS:
        return jsonify(
            {
                "error": f"Unsupported file type '{ext}'. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
            }
        ), 400

    file_bytes = file.read()
    try:
        rows = _get_rows(file_bytes, ext)
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400

    col_headers, parsed = _parse_rows(rows)
    if not parsed:
        return jsonify({"error": "CSV is empty or could not be parsed"}), 400

    db = get_nursing_employees_db()

    # Detect if we're in nursing mode
    raw_mode = request.headers.get("X-App-Mode", "").strip()
    mode_str = unquote(raw_mode).replace(" ", "_").replace("-", "_") if raw_mode else ""
    is_nursing = mode_str.startswith("nursing")

    # Optional: import into a specific department (overrides any dept column in CSV)
    target_department = request.args.get("department", "").strip() or None

    from .departments import build_department_list, auto_add_department
    known_departments = set(build_department_list())
    invalid_depts: set = set()
    for e in parsed:
        if target_department:
            e["home_department"] = target_department
        else:
            dept = e.get("home_department")
            if dept and dept not in known_departments:
                if is_nursing:
                    auto_add_department(dept)
                    known_departments.add(dept)
                else:
                    invalid_depts.add(dept)
                    e["home_department"] = None

    # ── Gender-based name disambiguation ─────────────────────────────────────
    # Detect which attribute columns represent gender
    male_attr = next((f"col_{i+1}" for i, h in enumerate(col_headers) if h == "גבר"), None)
    female_attr = next((f"col_{i+1}" for i, h in enumerate(col_headers) if h == "אישה"), None)

    def _gender(attrs):
        if male_attr and male_attr in attrs:
            return "male"
        if female_attr and female_attr in attrs:
            return "female"
        return None

    def _suffix(gender):
        return " בן" if gender == "male" else " בת" if gender == "female" else ""

    # Step 1: disambiguate name conflicts WITHIN the imported CSV
    from collections import defaultdict
    name_groups = defaultdict(list)
    for e in parsed:
        name_groups[e["name"]].append(e)

    for base_name, group in name_groups.items():
        if len(group) > 1:
            genders = [_gender(e["attributes"]) for e in group]
            if len(set(genders)) > 1:          # at least two different genders
                for e in group:
                    e["name"] = base_name + _suffix(_gender(e["attributes"]))

    # Step 2: resolve conflicts between imported employees and existing DB employees,
    # then upsert (update if name exists, insert if new).
    # For general imports (no target_department) also remove employees no longer in the file.
    imported_names = set()
    for e in parsed:
        base_name = e["name"]
        existing = db.employees.find_one({"name": base_name})

        if existing:
            existing_gender = _gender(existing.get("attributes", []))
            import_gender   = _gender(e["attributes"])

            if existing_gender and import_gender and existing_gender != import_gender:
                # Same name, different gender — rename the existing employee and add suffix to new
                existing_suffix = _suffix(existing_gender)
                db.employees.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {"name": base_name + existing_suffix}}
                )
                e["name"] = base_name + _suffix(import_gender)
                # Insert the newly-suffixed employee
                doc = {"name": e["name"], "attributes": e["attributes"]}
                if e.get("home_department"):
                    doc["home_department"] = e["home_department"]
                db.employees.insert_one(doc)
            else:
                # Same name, same (or unknown) gender — update in place
                patch = {"attributes": e["attributes"]}
                if e.get("home_department") is not None:
                    patch["home_department"] = e["home_department"]
                elif "home_department" in e:
                    patch["home_department"] = None
                db.employees.update_one({"_id": existing["_id"]}, {"$set": patch})
        else:
            doc = {"name": e["name"], "attributes": e["attributes"]}
            if e.get("home_department"):
                doc["home_department"] = e["home_department"]
            db.employees.insert_one(doc)

        imported_names.add(e["name"])

    # Department-specific imports never delete employees from other departments.
    # General imports also do not delete employees missing from the file —
    # import is additive/update-only; deletions must be done manually.

    # Store column headers for UI display
    if col_headers:
        db.config.replace_one(
            {"key": "csv_column_headers"},
            {"key": "csv_column_headers", "headers": col_headers},
            upsert=True,
        )

    employees = [_serialize(e) for e in db.employees.find()]

    # Auto-seed default shift types on first employee import (if collection is empty)
    auto_seeded = False
    if db.shift_types.count_documents({}) == 0:
        csv_path = os.path.normpath(_DEFAULTS_CSV)
        if os.path.exists(csv_path):
            from .shift_types import _parse_shift_types_csv, _header_to_attr_map

            with open(csv_path, encoding="utf-8-sig") as f:
                text = f.read()
            parsed_shifts, _ = _parse_shift_types_csv(text, _header_to_attr_map(db))
            if parsed_shifts:
                db.shift_types.insert_many(parsed_shifts)
                auto_seeded = True

    # Nursing: auto-provision a per-department DB for every unique department
    if is_nursing:
        all_depts = {e.get("home_department") for e in parsed if e.get("home_department")}
        for dept in all_depts:
            ensure_nursing_dept_db(dept)

    return jsonify(
        {
            "imported": len(employees),
            "employees": employees,
            "column_headers": col_headers,
            "shift_types_auto_seeded": auto_seeded,
            "new_departments": [],
            "invalid_departments": sorted(invalid_depts),
        }
    ), 201


@employees_bp.get("/employees/column-headers")
def get_column_headers():
    db = get_nursing_employees_db()
    config = db.config.find_one({"key": "csv_column_headers"})
    headers = config["headers"] if config else []
    return jsonify(headers)


@employees_bp.patch("/employees/column-headers")
def rename_column_header():
    """Rename a single column header. Body: {index: 0-based int, name: str}"""
    db = get_nursing_employees_db()
    data = request.get_json()
    idx = data.get("index")
    new_name = (data.get("name") or "").strip()
    if idx is None or not new_name:
        return jsonify({"error": "index and name are required"}), 400
    config = db.config.find_one({"key": "csv_column_headers"})
    headers = list(config["headers"]) if config else []
    if idx < 0 or idx >= len(headers):
        return jsonify({"error": "column index out of range"}), 400
    headers[idx] = new_name
    db.config.update_one({"key": "csv_column_headers"}, {"$set": {"headers": headers}})
    return jsonify(headers)


@employees_bp.post("/employees/column-headers")
def add_column_header():
    """Append a new column header. Body: {name: str}"""
    db = get_nursing_employees_db()
    data = request.get_json()
    new_name = (data.get("name") or "").strip()
    if not new_name:
        return jsonify({"error": "name is required"}), 400
    config = db.config.find_one({"key": "csv_column_headers"})
    headers = list(config["headers"]) if config else []
    headers.append(new_name)
    db.config.update_one(
        {"key": "csv_column_headers"},
        {"$set": {"headers": headers}},
        upsert=True,
    )
    return jsonify(headers), 201


_ALL_EMPLOYEE_NAMES = [
    "ישראל ישראלי", "דוד כהן", "שרה לוי", "מיכל אברהם",
    "יוסי בן דוד", "רחל מזרחי", "אורית פרידמן", "עמית שפירא",
    "נועה ברק", "אבי גורן", "תמר שמש", "גיל אביב",
    "ליאור כץ", "רוני פרץ", "הילה גלעד", "מור זהבי",
    "אלון ברק", "שירה ניר", "עידו מלכה", "ורד שלום",
    "בני אוחיון", "קרן ביטון", "אייל דהן", "ספיר חזן",
    "ירון לוינסון", "מיה רוזן", "שי גרינברג", "עינב מנחם",
    "ניר שגיא", "יעל פלד", "אסף וייס", "חן גבאי",
    "דן לוי", "רינת שפר", "אמיר זיו", "תהל אשר",
    "אור בן-דוד", "סיון פינטו", "יונתן שרון", "מאיה עמית",
]


# Default nursing staff attribute columns (in order → col_1 … col_N)
_NURSING_DEFAULT_ATTRIBUTES = [
    "אח/אחות",        # col_1 — base nurse role
    "אחראי משמרת",    # col_2 — shift supervisor (sub-role of nurse)
    "על בסיסי",       # col_3 — permanently stationed (sub-attr of supervisor)
    "גבר",            # col_4 — male gender
    "אישה",           # col_5 — female gender
    "ותיק",           # col_6 — veteran
    "צעיר",           # col_7 — junior
    "כוח עזר",        # col_8 — aide (not a nurse)
    "סטודנט",         # col_9 — student (not nurse, not aide)
]


@employees_bp.post("/employees/seed-nursing-attributes")
def seed_nursing_attributes():
    """
    Seed the default nursing attribute column headers.
    Additive: only appends headers that do not already exist.
    Returns the updated headers list.
    """
    db = get_nursing_employees_db()
    config = db.config.find_one({"key": "csv_column_headers"})
    existing = list(config["headers"]) if config else []
    existing_set = set(existing)

    added = []
    for attr in _NURSING_DEFAULT_ATTRIBUTES:
        if attr not in existing_set:
            existing.append(attr)
            existing_set.add(attr)
            added.append(attr)

    db.config.update_one(
        {"key": "csv_column_headers"},
        {"$set": {"headers": existing}},
        upsert=True,
    )
    return jsonify({"ok": True, "headers": existing, "added": added}), 200


@employees_bp.post("/employees/seed-defaults")
def seed_default_employees():
    """Seed example employees (unique per department) if none exist."""
    db = get_nursing_employees_db()
    if db.employees.count_documents({}) > 0:
        return jsonify({"ok": False, "message": "employees already exist"}), 409

    # Derive a stable offset from the database name so each department gets different names
    db_name = db.name  # e.g. "scheduler_הדס"
    offset = sum(ord(c) for c in db_name) % len(_ALL_EMPLOYEE_NAMES)
    names = [_ALL_EMPLOYEE_NAMES[(offset + i) % len(_ALL_EMPLOYEE_NAMES)] for i in range(10)]

    to_insert = [
        {"name": n, "attributes": [], "max_shifts_per_week": 6}
        for n in names
    ]
    db.employees.insert_many(to_insert)
    result = [_serialize(e) for e in db.employees.find()]
    return jsonify({"ok": True, "seeded": len(result), "employees": result}), 201


@employees_bp.delete("/employees/column-headers/<int:col_index>")
def delete_column_header(col_index):
    """
    Delete a column header by 0-based index.
    Removes the corresponding col_N attribute from all employees.
    All col indices > deleted index are shifted down by 1 on affected employees.
    """
    db = get_nursing_employees_db()
    config = db.config.find_one({"key": "csv_column_headers"})
    headers = list(config["headers"]) if config else []
    if col_index < 0 or col_index >= len(headers):
        return jsonify({"error": "column index out of range"}), 400

    # col_index is 0-based in the UI / config, but col_attr uses 1-based: col_1, col_2, …
    deleted_attr = col_attr(col_index + 1)

    # Remove deleted attribute and shift down all higher-indexed col_N attributes
    for emp in db.employees.find():
        attrs = emp.get("attributes", [])
        new_attrs = []
        changed = False
        for a in attrs:
            if a == deleted_attr:
                changed = True  # drop it
            elif a.startswith("col_"):
                try:
                    n = int(a[4:])
                except ValueError:
                    new_attrs.append(a)
                    continue
                if n > col_index + 1:
                    new_attrs.append(col_attr(n - 1))
                    changed = True
                else:
                    new_attrs.append(a)
            else:
                new_attrs.append(a)
        if changed:
            db.employees.update_one(
                {"_id": emp["_id"]}, {"$set": {"attributes": new_attrs}}
            )

    # Also shift required_attributes in shift_types
    for st in db.shift_types.find():
        req = st.get("required_attributes", [])
        new_req = []
        changed = False
        for a in req:
            if a == deleted_attr:
                changed = True
            elif a.startswith("col_"):
                try:
                    n = int(a[4:])
                except ValueError:
                    new_req.append(a)
                    continue
                if n > col_index + 1:
                    new_req.append(col_attr(n - 1))
                    changed = True
                else:
                    new_req.append(a)
            else:
                new_req.append(a)
        if changed:
            db.shift_types.update_one(
                {"_id": st["_id"]}, {"$set": {"required_attributes": new_req}}
            )

    # Remove the header and persist
    headers.pop(col_index)
    db.config.update_one({"key": "csv_column_headers"}, {"$set": {"headers": headers}})

    return jsonify(headers)
